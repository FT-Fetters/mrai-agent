import os
from mrai.agent.schema import Tool
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from openpyxl.cell import Cell
from openpyxl.utils.cell import coordinate_from_string
from typing import Optional, Tuple, List, Any

def excel_tool_list():
    """返回excel相关的工具列表"""
    return [
        ReadExcelBaseInfoTool(),
        ReadCellTool(),
        WriteCellTool(),
        WriteCellRangeTool(),
        CreateExcelTool(),
        FormatCellRangeTool(),
        MergeCellsTool(),
        UnmergeCellsTool(),
        InsertRowsTool(),
        InsertColsTool(),
        DeleteRowsTool(),
        DeleteColsTool()
    ]


class ReadExcelBaseInfoTool(Tool):

    def __init__(self):
        super().__init__(
            name="read_excel_base_info",
            description="read the base info of the excel file",
            parameters={
                "file_path": Tool.ToolParameter(
                    name="file_path",
                    description="The path to the excel file",
                    type="string",
                    required=True
                )
            }
        )
    
    def execute(self, file_path: str):
        
        # check if the file exists
        if not os.path.exists(file_path):
            return f"Error: file {file_path} not found"
        
        try:
            # open the excel workbook
            wb = load_workbook(file_path, data_only=True)
            
            # get the sheet info
            sheet_names = wb.sheetnames
            active_sheet = wb.active.title if wb.active else None
            
            # build the workbook info
            workbook_info = (
                f"file name: {os.path.basename(file_path)}\n"
                f"sheet number: {len(sheet_names)}\n"
                f"sheet list: {', '.join(sheet_names)}\n"
                f"current active sheet: {active_sheet}\n"
            )
            
            return workbook_info
        except Exception as e:
            return f"Error: {str(e)}"


class ReadCellTool(Tool):

    def __init__(self):
        super().__init__(
            name="read_cell",
            description="read the cell value of the excel file",
            parameters={
                "file_path": Tool.ToolParameter(
                    name="file_path",
                    description="The path to the excel file",
                    type="string",
                    required=True
                ),
                "sheet_name": Tool.ToolParameter(
                    name="sheet_name",
                    description="The name of the sheet to read",
                    type="string",
                    required=True
                ),
                "start_cell": Tool.ToolParameter(
                    name="start_cell",
                    description="The start cell of the range to read, if not provided, the whole sheet will be read",
                    type="string",
                    required=False
                ),
                "end_cell": Tool.ToolParameter(
                    name="end_cell",
                    description="The end cell of the range to read, if not provided, the whole sheet will be read",
                    type="string",
                    required=False
                )
            }
        )
    
    def _read_excel_data(self, sheet, start_cell: Optional[str], end_cell: Optional[str]) -> Tuple[List[List[Any]], int, int, int, int]:
        """Reads data from the sheet based on cell range and returns data with coordinates."""
        data = []
        start_row_idx = 1
        start_col_idx = 1
        num_rows = 0
        num_cols = 0

        if start_cell and end_cell:
            # Read specific range
            try:
                col_letter_start, row_num_start = coordinate_from_string(start_cell)
                start_row_idx = row_num_start
                start_col_idx = column_index_from_string(col_letter_start)
                cell_range = sheet[f"{start_cell}:{end_cell}"]
            except ValueError:
                raise ValueError(f"Invalid cell coordinate in range '{start_cell}:{end_cell}'")

            if isinstance(cell_range, Cell):
                data = [[cell_range.value]]
            else:
                # Handle multi-cell/single-row range
                if isinstance(cell_range, tuple):
                    if cell_range and isinstance(cell_range[0], tuple): # Multi-row
                        for row_tuple in cell_range:
                            data.append([cell.value for cell in row_tuple])
                    else: # Single row
                        data.append([cell.value for cell in cell_range])

            num_rows = len(data)
            num_cols = max(len(row) for row in data) if data else 0

        elif start_cell:
            # Read single cell
            try:
                col_letter, row_num = coordinate_from_string(start_cell)
                start_row_idx = row_num
                start_col_idx = column_index_from_string(col_letter)
                data = [[sheet[start_cell].value]]
                num_rows = 1
                num_cols = 1
            except ValueError:
                 raise ValueError(f"Invalid cell coordinate '{start_cell}'")
        else:
            # Read entire used range
            if not sheet.calculate_dimension(): # Handle truly empty sheet
                 return [], 1, 1, 0, 0
            min_row, min_col, max_row, max_col = range_boundaries(sheet.calculate_dimension())

            if min_row is None: # Another check for empty sheet after dimension calculation
                return [], 1, 1, 0, 0

            start_row_idx = min_row
            # Ensure min_col is not None before assigning to start_col_idx (which expects int)
            start_col_idx = min_col if min_col is not None else 1

            # Iterate within the calculated bounds
            data = []
            # Adjust iteration if min_col was None (shouldn't happen if min_row wasn't None, but safer)
            iter_min_col = min_col if min_col is not None else 1
            iter_max_col = max_col if max_col is not None else 1
            for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=iter_min_col, max_col=iter_max_col, values_only=True):
                data.append(list(row))

            # Filter out fully empty rows *after* reading the block to maintain relative indices correctly for start_row_idx
            # Note: This approach might include rows of None within the data block if they exist.
            # A more complex approach would be needed to find the absolute minimal bounding box containing non-None values.
            # For simplicity, we stick to the dimension reported by openpyxl.

            num_rows = len(data) # max_row - min_row + 1
            num_cols = len(data[0]) if num_rows > 0 else 0 # max_col - min_col + 1

        # Final check on dimensions based on read data
        if data:
            actual_num_rows = len(data)
            # Recalculate max columns based on actual data fetched
            actual_num_cols = max(len(r) for r in data) if actual_num_rows > 0 else 0
            num_rows = actual_num_rows
            # Make sure num_cols reflects the actual data width, not just range_boundaries width
            # Also handle case where start_col_idx was 1 but iter_min_col might have been different
            # The most reliable num_cols comes from the data itself
            num_cols = actual_num_cols
        else:
             num_rows = 0
             num_cols = 0

        # Ensure the returned start_col_idx is always an int
        final_start_col = start_col_idx if start_col_idx is not None else 1

        return data, start_row_idx, final_start_col, num_rows, num_cols

    def _format_data_as_table(self, data: List[List[Any]], start_row_idx: int, start_col_idx: int, num_rows: int, num_cols: int) -> str:
        """Formats the data list into a string table with coordinates."""
        if not data or num_rows == 0 or num_cols == 0:
            return "No data found or sheet is empty." # Or tailor message based on initial context if needed

        # Convert all data to string, handle None
        str_data = [[str(val) if val is not None else "" for val in row] for row in data]

        # Pad rows to ensure all have num_cols
        padded_data = []
        for row in str_data:
            padded_row = row + [""] * (num_cols - len(row))
            padded_data.append(padded_row)
        str_data = padded_data

        # Calculate max width for each column
        max_widths = [0] * num_cols
        for row in str_data:
            for i, cell in enumerate(row):
                 if i < num_cols:
                    max_widths[i] = max(max_widths[i], len(cell))

        # Generate row headers (numbers)
        row_headers = [str(r) for r in range(start_row_idx, start_row_idx + num_rows)]
        row_header_width = max(len(h) for h in row_headers) if row_headers else 0

        # Generate column headers (letters)
        col_headers_raw = [get_column_letter(c) for c in range(start_col_idx, start_col_idx + num_cols)]
        col_headers_padded = [col_headers_raw[j].ljust(max_widths[j]) for j in range(num_cols)]

        # Build the formatted string list
        formatted_lines = []

        # Top-left corner space + Column Headers
        top_left_corner = " " * (row_header_width + 1)
        column_header_line = top_left_corner + "| " + " | ".join(col_headers_padded) + " |"
        formatted_lines.append(column_header_line)

        # Separator line
        separator_data_part = "-+-" .join("-" * max_widths[j] for j in range(num_cols))
        separator_line = "-" * row_header_width + "-+" + separator_data_part + "-+"
        formatted_lines.append(separator_line)

        # Data rows with row headers
        for i, row in enumerate(str_data):
            if i < len(row_headers):
                row_header = row_headers[i].rjust(row_header_width)
                formatted_cells = [cell.ljust(max_widths[j]) for j, cell in enumerate(row)]
                formatted_row = row_header + " | " + " | ".join(formatted_cells) + " |"
                formatted_lines.append(formatted_row)

        return "\\n".join(formatted_lines)

    def execute(self, file_path: str, sheet_name: str, start_cell: Optional[str] = None, end_cell: Optional[str] = None):
        if not os.path.exists(file_path):
            return f"Error: file {file_path} not found"

        wb = None
        try:
            wb = load_workbook(filename=file_path, data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames:
                return f"Error: sheet {sheet_name} not found in file {file_path}"
            sheet = wb[sheet_name]

            try:
                 # Delegate data reading and coordinate finding
                 data, start_row, start_col, num_rows, num_cols = self._read_excel_data(sheet, start_cell, end_cell)

                 # Check if data was found before formatting
                 if not data or num_rows == 0 or num_cols == 0:
                      # Provide context-specific messages based on input
                      if start_cell or end_cell:
                           return "No data found in the specified range."
                      else:
                           return "The sheet is empty."

                 # Delegate formatting
                 return self._format_data_as_table(data, start_row, start_col, num_rows, num_cols)

            except ValueError as ve:
                 # Catch specific errors from _read_excel_data (e.g., invalid coordinates)
                 return f"Error: {str(ve)}"

        except Exception as e:
            # import traceback
            # traceback.print_exc() # Uncomment for detailed debugging
            return f"Error reading Excel file: {str(e)}"
        finally:
            if wb:
                wb.close()


class WriteCellTool(Tool):
    

    def __init__(self):
        super().__init__(
            name="write_cell",
            description="write the cell value of the excel file",
            parameters={
                "file_path": Tool.ToolParameter(
                    name="file_path",
                    description="The path to the excel file",
                    type="string",
                    required=True
                ),
                "sheet_name": Tool.ToolParameter(
                    name="sheet_name",
                    description="The name of the sheet to write",
                    type="string",
                    required=True
                ),
                "cell": Tool.ToolParameter(
                    name="cell",
                    description="The cell to write, e.g. A1",
                    type="string",
                    required=True
                ),
                "value": Tool.ToolParameter(
                    name="value",
                    description="The value to write, e.g. 123",
                    type="string",
                    required=True
                )
            }
        )
    
    def execute(self, file_path: str, sheet_name: str, cell: str, value: str):
        wb = None
        try:
            # Check if the file exists, if not, create a new workbook
            if os.path.exists(file_path):
                wb = load_workbook(filename=file_path)
            else:
                wb = Workbook()
                # Remove the default 'Sheet' if we are creating a new sheet with a specific name
                if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1:
                    default_sheet = wb['Sheet']
                    wb.remove(default_sheet)


            # Get or create the sheet
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
            else:
                sheet = wb.create_sheet(title=sheet_name)
                # If it was a new workbook, make this the active sheet
                if len(wb.sheetnames) == 1:
                   wb.active = sheet


            # Write the value to the cell
            try:
                sheet[cell] = value
            except ValueError:
                 return f"Error: Invalid cell coordinate '{cell}'"


            # Save the workbook
            wb.save(file_path)
            
            # Close the workbook
            wb.close()
            
            return f"Successfully wrote '{value}' to cell {cell} in sheet '{sheet_name}' of file '{file_path}'"


        except Exception as e:
            # import traceback
            # traceback.print_exc() # Uncomment for detailed debugging
            return f"Error writing to Excel file: {str(e)}"
        finally:
             if wb:
                try:
                    wb.close() # Close workbook if it was loaded (not strictly necessary for write, but good practice)
                except Exception:
                    pass # Ignore potential errors on close if save failed etc.


class WriteCellRangeTool(Tool):

    def __init__(self):
        super().__init__(
            name="write_cell_range",
            description="write the cell value of the excel file",
            parameters={
                "file_path": Tool.ToolParameter(
                    name="file_path",
                    description="The path to the excel file",
                    type="string",
                    required=True
                ),
                "sheet_name": Tool.ToolParameter(
                    name="sheet_name",
                    description="The name of the sheet to write",
                    type="string",
                    required=True
                ),
                "start_cell": Tool.ToolParameter(
                    name="start_cell",
                    description="The start cell of the range to write",
                    type="string",
                    required=True
                ),
                "end_cell": Tool.ToolParameter(
                    name="end_cell",
                    description="The end cell of the range to write",
                    type="string",
                    required=True
                ),
                "data": Tool.ToolParameter(
                    name="data",
                    description="""The data to write, the data should be a list of lists, e.g. [["a", "b", "c"], [1, 2, 3]]""",
                    type="list",
                    required=True
                )
            }
        )
    
    def execute(self, file_path: str, sheet_name: str, start_cell: str, end_cell: str, data: List[List[Any]]):
        wb = None
        try:
            # Check if the file exists, if not, create a new workbook
            if os.path.exists(file_path):
                wb = load_workbook(filename=file_path)
            else:
                wb = Workbook()
                # Remove the default 'Sheet' if we are creating a new sheet with a specific name
                if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1:
                    default_sheet = wb['Sheet']
                    wb.remove(default_sheet)


            # Get or create the sheet
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
            else:
                sheet = wb.create_sheet(title=sheet_name)
                # If it was a new workbook, make this the active sheet
                if len(wb.sheetnames) == 1:
                    wb.active = sheet

            # Parse start cell
            try:
                start_col_letter, start_row_num = coordinate_from_string(start_cell)
                start_col_idx = column_index_from_string(start_col_letter)
            except ValueError:
                return f"Error: Invalid start cell coordinate '{start_cell}'"

            # Write data
            if not isinstance(data, list) or not all(isinstance(row, list) for row in data):
                 return "Error: Invalid data format. 'data' should be a list of lists."
            
            for r_idx, row_data in enumerate(data):
                for c_idx, value in enumerate(row_data):
                    target_row = start_row_num + r_idx
                    target_col = start_col_idx + c_idx
                    try:
                        sheet.cell(row=target_row, column=target_col, value=value)
                    except Exception as cell_write_error:
                         # More specific error handling could be added here if needed
                         return f"Error writing value '{value}' to cell at row {target_row}, column {target_col}: {cell_write_error}"

            # Calculate end cell based on data dimensions for the success message
            if data and data[0]:
                end_row_num = start_row_num + len(data) - 1
                end_col_idx = start_col_idx + len(data[0]) - 1
                calculated_end_cell = f"{get_column_letter(end_col_idx)}{end_row_num}"
            else:
                calculated_end_cell = start_cell # No data written, range is just the start cell

            # Save the workbook
            wb.save(file_path)
            return f"Successfully wrote data to range {start_cell}:{calculated_end_cell} in sheet '{sheet_name}' of file '{file_path}'"

        except Exception as e:
            # import traceback
            # traceback.print_exc() # Uncomment for detailed debugging
            return f"Error writing range to Excel file: {str(e)}"
        finally:
             if wb:
                try:
                    wb.close()
                except Exception:
                    pass # Ignore potential errors on close


class CreateExcelTool(Tool):

    def __init__(self):
        super().__init__(
            name="create_excel",
            description="create a new excel file",
            parameters={
                "file_path": Tool.ToolParameter(
                    name="file_path",
                    description="The path to the excel file",
                    type="string",
                    required=True
                ),
                "sheet_name": Tool.ToolParameter(
                    name="sheet_name",
                    description="The name of the first sheet (default: 'Sheet')",
                    type="string",
                    required=False
                )
            }
        )
    
    def execute(self, file_path: str, sheet_name: Optional[str] = None):
        wb = None
        try:
            # 检查文件目录是否存在，如果不存在则创建
            file_dir = os.path.dirname(file_path)
            if file_dir and not os.path.exists(file_dir):
                os.makedirs(file_dir)
                
            # 检查文件是否已存在
            if os.path.exists(file_path):
                return f"Error: file {file_path} already exists. Please use another file path or delete the existing file."
                
            # 创建工作簿
            wb = Workbook()
            
            # 如果提供了自定义的sheet名称，重命名默认sheet
            if sheet_name:
                default_sheet = wb.active
                if default_sheet:
                    default_sheet.title = sheet_name
                
            # 保存工作簿
            wb.save(file_path)
            
            # 返回成功信息
            sheet_info = f" with sheet '{sheet_name}'" if sheet_name else ""
            return f"Successfully created excel file at {file_path}{sheet_info}"
            
        except PermissionError:
            return f"Error: Permission denied when creating file at {file_path}"
        except Exception as e:
            return f"Error creating Excel file: {str(e)}"
        finally:
            if wb:
                try:
                    wb.close()
                except Exception:
                    pass # Ignore potential errors on close


class FormatCellRangeTool(Tool):
    def __init__(self):
        super().__init__(
            name="format_cell_range",
            description="Sets the format for a range of cells in an Excel file.",
            parameters={
                "file_path": Tool.ToolParameter(
                    name="file_path", description="Path to the Excel file.", type="string", required=True
                ),
                "sheet_name": Tool.ToolParameter(
                    name="sheet_name", description="Name of the sheet to modify.", type="string", required=True
                ),
                "start_cell": Tool.ToolParameter(
                    name="start_cell", description="The top-left cell of the range to format (e.g., 'A1').", type="string", required=True
                ),
                "end_cell": Tool.ToolParameter(
                    name="end_cell", description="The bottom-right cell of the range to format (e.g., 'C3').", type="string", required=True
                ),
                "font_name": Tool.ToolParameter(
                    name="font_name", description="Font name (e.g., 'Arial').", type="string", required=False
                ),
                "font_size": Tool.ToolParameter(
                    name="font_size", description="Font size (e.g., 12).", type="number", required=False
                ),
                "bold": Tool.ToolParameter(
                    name="bold", description="Set font to bold (true/false).", type="boolean", required=False
                ),
                "italic": Tool.ToolParameter(
                    name="italic", description="Set font to italic (true/false).", type="boolean", required=False
                ),
                "underline": Tool.ToolParameter(
                    name="underline", description="Set font underline ('single', 'double', 'singleAccounting', 'doubleAccounting', None).", type="string", required=False
                ),
                "font_color": Tool.ToolParameter(
                    name="font_color", description="Font color as hex code (e.g., 'FF0000' for red).", type="string", required=False
                ),
                "background_color": Tool.ToolParameter(
                    name="background_color", description="Cell background color as hex code (e.g., 'FFFF00' for yellow).", type="string", required=False
                ),
                "horizontal_alignment": Tool.ToolParameter(
                    name="horizontal_alignment", description="Horizontal alignment ('general', 'left', 'center', 'right', 'fill', 'justify', 'centerContinuous', 'distributed').", type="string", required=False
                ),
                "vertical_alignment": Tool.ToolParameter(
                    name="vertical_alignment", description="Vertical alignment ('top', 'center', 'bottom', 'justify', 'distributed').", type="string", required=False
                ),
            }
        )

    def execute(self, file_path: str, sheet_name: str, start_cell: str, end_cell: str,
                font_name: Optional[str] = None, font_size: Optional[int] = None,
                bold: Optional[bool] = None, italic: Optional[bool] = None, underline: Optional[str] = None,
                font_color: Optional[str] = None, background_color: Optional[str] = None,
                horizontal_alignment: Optional[str] = None, vertical_alignment: Optional[str] = None):
        wb = None
        try:
            if not os.path.exists(file_path):
                return f"Error: file {file_path} not found"

            wb = load_workbook(filename=file_path)
            if sheet_name not in wb.sheetnames:
                return f"Error: sheet {sheet_name} not found in file {file_path}"
            sheet = wb[sheet_name]

            # --- Parse Range ---
            try:
                start_col_letter, start_row_num = coordinate_from_string(start_cell)
                end_col_letter, end_row_num = coordinate_from_string(end_cell)
                start_col_idx = column_index_from_string(start_col_letter)
                end_col_idx = column_index_from_string(end_col_letter)
                
                if start_row_num > end_row_num or start_col_idx > end_col_idx:
                     return f"Error: Invalid range order. Start cell {start_cell} must be top-left of end cell {end_cell}."

            except ValueError:
                return f"Error: Invalid cell coordinates in range '{start_cell}:{end_cell}'"
            except Exception as e:
                 return f"Error parsing range '{start_cell}:{end_cell}': {str(e)}"

            # --- Iterate and Apply Formatting ---
            for row_idx in range(start_row_num, end_row_num + 1):
                 for col_idx in range(start_col_idx, end_col_idx + 1):
                    try:
                        target_cell = sheet.cell(row=row_idx, column=col_idx)

                        # --- Apply Font Formatting ---
                        current_font = target_cell.font
                        new_font = Font(name=font_name if font_name is not None else current_font.name,
                                      size=font_size if font_size is not None else current_font.size,
                                      bold=bold if bold is not None else current_font.bold, # Apply if specified
                                      italic=italic if italic is not None else current_font.italic, # Apply if specified
                                      underline=underline if underline is not None else current_font.underline, # type: ignore
                                      color=font_color if font_color is not None else current_font.color)
                        target_cell.font = new_font

                        # --- Apply Background Fill ---
                        if background_color:
                            fill = PatternFill(start_color=background_color, end_color=background_color, fill_type="solid")
                            target_cell.fill = fill

                        # --- Apply Alignment ---
                        current_alignment = target_cell.alignment
                        new_alignment = Alignment(horizontal=horizontal_alignment if horizontal_alignment is not None else current_alignment.horizontal,
                                                vertical=vertical_alignment if vertical_alignment is not None else current_alignment.vertical,
                                                text_rotation=current_alignment.text_rotation,
                                                wrap_text=current_alignment.wrap_text,
                                                shrink_to_fit=current_alignment.shrink_to_fit,
                                                indent=current_alignment.indent)
                        target_cell.alignment = new_alignment

                    except Exception as cell_error:
                         cell_coord = f"{get_column_letter(col_idx)}{row_idx}"
                         return f"Error formatting cell {cell_coord}: {str(cell_error)}"

            wb.save(file_path)
            return f"Successfully formatted cell range {start_cell}:{end_cell} in sheet '{sheet_name}' of file '{file_path}'"

        except Exception as e:
            return f"Error formatting cell range: {str(e)}"
        finally:
            if wb:
                try:
                    wb.close()
                except Exception:
                    pass


class MergeCellsTool(Tool):
    def __init__(self):
        super().__init__(
            name="merge_cells",
            description="Merges one or more ranges of cells in an Excel sheet.",
            parameters={
                "file_path": Tool.ToolParameter(
                    name="file_path", description="Path to the Excel file.", type="string", required=True
                ),
                "sheet_name": Tool.ToolParameter(
                    name="sheet_name", description="Name of the sheet to modify.", type="string", required=True
                ),
                "ranges": Tool.ToolParameter(
                    name="ranges", description="A list of cell ranges to merge (e.g., ['A1:B2', 'D1:E2']).", type="list", required=True
                ),
            }
        )

    def execute(self, file_path: str, sheet_name: str, ranges: List[str]):
        wb = None
        results = {"success": [], "errors": []}
        try:
            if not os.path.exists(file_path):
                return f"Error: file {file_path} not found"

            wb = load_workbook(filename=file_path)
            if sheet_name not in wb.sheetnames:
                return f"Error: sheet {sheet_name} not found in file {file_path}"
            sheet = wb[sheet_name]

            if not isinstance(ranges, list):
                return "Error: 'ranges' parameter must be a list of range strings."

            for merge_range in ranges:
                if not isinstance(merge_range, str):
                    results["errors"].append(f"Invalid item in ranges list (must be string): {merge_range}")
                    continue
                try:
                    # Basic validation before attempting merge
                    if ':' not in merge_range:
                         raise ValueError("Range must contain ':'")
                    start_cell, end_cell = merge_range.split(':', 1)
                    coordinate_from_string(start_cell)
                    coordinate_from_string(end_cell)
                    
                    sheet.merge_cells(merge_range)
                    results["success"].append(merge_range)
                except ValueError as ve:
                    results["errors"].append(f"Invalid coordinates in range '{merge_range}': {str(ve)}")
                except Exception as e:
                    # Consider logging full traceback for debugging
                    results["errors"].append(f"Error merging range '{merge_range}': {str(e)}")

            if results["success"] or results["errors"]:
                wb.save(file_path)

            # --- Format output message ---
            output_message = f"Merge operation summary for sheet '{sheet_name}' in '{file_path}':\n"
            if results["success"]:
                output_message += f"Successfully merged ranges: {', '.join(results['success'])}\n"
            if results["errors"]:
                output_message += f"Errors encountered: {'; '.join(results['errors'])}"
            
            # If only errors occurred, prepend "Error: " to the message
            if not results["success"] and results["errors"]:
                 output_message = "Error: " + output_message
            elif not results["success"] and not results["errors"]:
                 output_message = "No valid ranges provided to merge."
                 
            return output_message.strip()

        except Exception as e:
            return f"Error during merge operation: {str(e)}"
        finally:
            if wb:
                try:
                    wb.close()
                except Exception:
                    pass


class UnmergeCellsTool(Tool):
    def __init__(self):
        super().__init__(
            name="unmerge_cells",
            description="Unmerges one or more previously merged ranges of cells in an Excel sheet.",
             parameters={
                "file_path": Tool.ToolParameter(
                    name="file_path", description="Path to the Excel file.", type="string", required=True
                ),
                "sheet_name": Tool.ToolParameter(
                    name="sheet_name", description="Name of the sheet to modify.", type="string", required=True
                ),
                "ranges": Tool.ToolParameter(
                    name="ranges", description="A list of cell ranges to unmerge (e.g., ['A1:B2', 'D1:E2']).", type="list", required=True
                ),
            }
        )

    def execute(self, file_path: str, sheet_name: str, ranges: List[str]):
        wb = None
        results = {"success": [], "errors": []}
        try:
            if not os.path.exists(file_path):
                return f"Error: file {file_path} not found"

            wb = load_workbook(filename=file_path)
            if sheet_name not in wb.sheetnames:
                return f"Error: sheet {sheet_name} not found in file {file_path}"
            sheet = wb[sheet_name]

            if not isinstance(ranges, list):
                return "Error: 'ranges' parameter must be a list of range strings."

            for unmerge_range in ranges:
                if not isinstance(unmerge_range, str):
                     results["errors"].append(f"Invalid item in ranges list (must be string): {unmerge_range}")
                     continue
                try:
                    # Basic validation before attempting unmerge
                    if ':' not in unmerge_range:
                        raise ValueError("Range must contain ':'")
                    start_cell, end_cell = unmerge_range.split(':', 1)
                    coordinate_from_string(start_cell)
                    coordinate_from_string(end_cell)
                    
                    sheet.unmerge_cells(unmerge_range)
                    results["success"].append(unmerge_range)
                except ValueError as ve:
                    results["errors"].append(f"Invalid coordinates in range '{unmerge_range}': {str(ve)}")
                except Exception as e:
                    # openpyxl might raise errors if range wasn't merged, catch them
                    results["errors"].append(f"Error unmerging range '{unmerge_range}': {str(e)}")

            if results["success"] or results["errors"]:
                wb.save(file_path)

            # --- Format output message ---
            output_message = f"Unmerge operation summary for sheet '{sheet_name}' in '{file_path}':\n"
            if results["success"]:
                output_message += f"Successfully unmerged ranges: {', '.join(results['success'])}\n"
            if results["errors"]:
                output_message += f"Errors encountered: {'; '.join(results['errors'])}"
                
            # If only errors occurred, prepend "Error: "
            if not results["success"] and results["errors"]:
                output_message = "Error: " + output_message
            elif not results["success"] and not results["errors"]:
                 output_message = "No valid ranges provided to unmerge."

            return output_message.strip()

        except Exception as e:
            return f"Error during unmerge operation: {str(e)}"
        finally:
            if wb:
                try:
                    wb.close()
                except Exception:
                    pass


class InsertRowsTool(Tool):
    def __init__(self):
        super().__init__(
            name="insert_rows",
            description="Inserts one or more rows into an Excel sheet.",
            parameters={
                "file_path": Tool.ToolParameter(
                    name="file_path", description="Path to the Excel file.", type="string", required=True
                ),
                "sheet_name": Tool.ToolParameter(
                    name="sheet_name", description="Name of the sheet to modify.", type="string", required=True
                ),
                "index": Tool.ToolParameter(
                    name="index", description="The row number before which to insert rows (1-based).", type="number", required=True
                ),
                "amount": Tool.ToolParameter(
                    name="amount", description="The number of rows to insert (default is 1).", type="number", required=False
                ),
            }
        )

    def execute(self, file_path: str, sheet_name: str, index: int, amount: int = 1):
        wb = None
        try:
            if not os.path.exists(file_path):
                return f"Error: file {file_path} not found"
            if amount < 1:
                return "Error: Amount must be a positive integer."
            if index < 1:
                 return "Error: Row index must be 1 or greater."

            wb = load_workbook(filename=file_path)
            if sheet_name not in wb.sheetnames:
                return f"Error: sheet {sheet_name} not found in file {file_path}"
            sheet = wb[sheet_name]

            try:
                 sheet.insert_rows(idx=index, amount=amount)
            except Exception as e:
                 return f"Error inserting rows at index {index}: {str(e)}"

            wb.save(file_path)
            return f"Successfully inserted {amount} row(s) before row {index} in sheet '{sheet_name}' of file '{file_path}'"

        except Exception as e:
            return f"Error inserting rows: {str(e)}"
        finally:
            if wb:
                try:
                    wb.close()
                except Exception:
                    pass


class InsertColsTool(Tool):
    def __init__(self):
        super().__init__(
            name="insert_cols",
            description="Inserts one or more columns into an Excel sheet.",
            parameters={
                "file_path": Tool.ToolParameter(
                    name="file_path", description="Path to the Excel file.", type="string", required=True
                ),
                "sheet_name": Tool.ToolParameter(
                    name="sheet_name", description="Name of the sheet to modify.", type="string", required=True
                ),
                "index": Tool.ToolParameter(
                    name="index", description="The column letter (e.g., 'C') or number (e.g., 3 for 1-based) before which to insert columns.", type="string", required=True
                ),
                "amount": Tool.ToolParameter(
                    name="amount", description="The number of columns to insert (default is 1).", type="number", required=False
                ),
            }
        )

    def execute(self, file_path: str, sheet_name: str, index: Any, amount: int = 1):
        wb = None
        try:
            if not os.path.exists(file_path):
                return f"Error: file {file_path} not found"
            if amount < 1:
                return "Error: Amount must be a positive integer."

            # Convert column index if it's a letter or number-like string
            col_idx: int
            index_str = str(index) # Convert input to string for consistent processing
            if index_str.isalpha(): # Check if it's purely alphabetical (column letter)
                try:
                    col_idx = column_index_from_string(index_str.upper())
                except ValueError:
                    return f"Error: Invalid column letter '{index_str}'"
            elif index_str.isdigit(): # Check if it's purely digits (column number)
                 col_num = int(index_str)
                 if col_num < 1:
                     return "Error: Column index number must be 1 or greater."
                 col_idx = col_num
            else:
                 return f"Error: Invalid column index format '{index}'. Use letter (e.g., 'C') or positive number (e.g., 3)."


            wb = load_workbook(filename=file_path)
            if sheet_name not in wb.sheetnames:
                return f"Error: sheet {sheet_name} not found in file {file_path}"
            sheet = wb[sheet_name]

            try:
                 sheet.insert_cols(idx=col_idx, amount=amount)
            except Exception as e:
                 return f"Error inserting columns at index {col_idx}: {str(e)}"


            wb.save(file_path)
            original_index_str = get_column_letter(col_idx) # Use converted index for consistent message
            return f"Successfully inserted {amount} column(s) before column {original_index_str} (index {col_idx}) in sheet '{sheet_name}' of file '{file_path}'"

        except Exception as e:
            return f"Error inserting columns: {str(e)}"
        finally:
            if wb:
                try:
                    wb.close()
                except Exception:
                    pass


class DeleteRowsTool(Tool):
    def __init__(self):
        super().__init__(
            name="delete_rows",
            description="Deletes one or more rows from an Excel sheet.",
             parameters={
                "file_path": Tool.ToolParameter(
                    name="file_path", description="Path to the Excel file.", type="string", required=True
                ),
                "sheet_name": Tool.ToolParameter(
                    name="sheet_name", description="Name of the sheet to modify.", type="string", required=True
                ),
                "index": Tool.ToolParameter(
                    name="index", description="The starting row number to delete (1-based).", type="number", required=True
                ),
                "amount": Tool.ToolParameter(
                    name="amount", description="The number of rows to delete (default is 1).", type="number", required=False
                ),
            }
        )

    def execute(self, file_path: str, sheet_name: str, index: int, amount: int = 1):
        wb = None
        try:
            if not os.path.exists(file_path):
                return f"Error: file {file_path} not found"
            if amount < 1:
                return "Error: Amount must be a positive integer."
            if index < 1:
                return "Error: Row index must be 1 or greater."


            wb = load_workbook(filename=file_path)
            if sheet_name not in wb.sheetnames:
                return f"Error: sheet {sheet_name} not found in file {file_path}"
            sheet = wb[sheet_name]

            try:
                # Optional: Add check if index + amount exceeds max_row? openpyxl might handle it.
                 sheet.delete_rows(idx=index, amount=amount)
            except Exception as e:
                 return f"Error deleting rows starting at index {index}: {str(e)}"


            wb.save(file_path)
            return f"Successfully deleted {amount} row(s) starting from row {index} in sheet '{sheet_name}' of file '{file_path}'"

        except Exception as e:
            return f"Error deleting rows: {str(e)}"
        finally:
            if wb:
                try:
                    wb.close()
                except Exception:
                    pass


class DeleteColsTool(Tool):
    def __init__(self):
        super().__init__(
            name="delete_cols",
            description="Deletes one or more columns from an Excel sheet.",
             parameters={
                "file_path": Tool.ToolParameter(
                    name="file_path", description="Path to the Excel file.", type="string", required=True
                ),
                "sheet_name": Tool.ToolParameter(
                    name="sheet_name", description="Name of the sheet to modify.", type="string", required=True
                ),
                "index": Tool.ToolParameter(
                    name="index", description="The starting column letter (e.g., 'C') or number (e.g., 3 for 1-based) to delete.", type="string", required=True
                ),
                "amount": Tool.ToolParameter(
                    name="amount", description="The number of columns to delete (default is 1).", type="number", required=False
                ),
            }
        )

    def execute(self, file_path: str, sheet_name: str, index: Any, amount: int = 1):
        wb = None
        try:
            if not os.path.exists(file_path):
                return f"Error: file {file_path} not found"
            if amount < 1:
                return "Error: Amount must be a positive integer."

            # Convert column index if it's a letter or number-like string
            col_idx: int
            index_str = str(index) # Convert input to string for consistent processing
            if index_str.isalpha(): # Check if it's purely alphabetical (column letter)
                try:
                    col_idx = column_index_from_string(index_str.upper())
                except ValueError:
                    return f"Error: Invalid column letter '{index_str}'"
            elif index_str.isdigit(): # Check if it's purely digits (column number)
                 col_num = int(index_str)
                 if col_num < 1:
                     return "Error: Column index number must be 1 or greater."
                 col_idx = col_num
            else:
                 return f"Error: Invalid column index format '{index}'. Use letter (e.g., 'C') or positive number (e.g., 3)."


            wb = load_workbook(filename=file_path)
            if sheet_name not in wb.sheetnames:
                return f"Error: sheet {sheet_name} not found in file {file_path}"
            sheet = wb[sheet_name]

            try:
                # Optional: Add check if index + amount exceeds max_column? openpyxl might handle it.
                 sheet.delete_cols(idx=col_idx, amount=amount)
            except Exception as e:
                 return f"Error deleting columns starting at index {col_idx}: {str(e)}"

            wb.save(file_path)
            original_index_str = get_column_letter(col_idx) # Use converted index for consistent message
            return f"Successfully deleted {amount} column(s) starting from column {original_index_str} (index {col_idx}) in sheet '{sheet_name}' of file '{file_path}'"

        except Exception as e:
            return f"Error deleting columns: {str(e)}"
        finally:
            if wb:
                try:
                    wb.close()
                except Exception:
                    pass
